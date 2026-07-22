"""
Exportación a Excel DACO — con historial de pagos
"""
from datetime import datetime, timezone
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.deps import CurrentUser, DBDep
from app.models.invoice_models import Invoice, InvoiceStatus
from app.models.invoice_payment_model import InvoicePayment
from app.models.quote_models import Quote
from app.models.models import LegalEntity

router = APIRouter(prefix="/exportar", tags=["Exportar"])

COLOR_HEADER = "1E293B"
COLOR_SUBHEADER = "334155"
COLOR_GREEN = "10B981"
COLOR_AMBER = "F59E0B"
COLOR_RED = "EF4444"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT = "F8FAFC"
COLOR_GRAY = "F1F5F9"


def cell_style(ws, row, col, value, bold=False, color=None, bg=None, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or "000000", name="Calibri", size=10)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        cell.number_format = number_format
    return cell


def add_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color="E2E8F0")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


@router.get("/estado-cuenta")
async def export_estado_cuenta(
    db: DBDep,
    current_user: CurrentUser,
    client_id: Optional[str] = Query(default=None),
):
    if client_id:
        clients_r = await db.execute(select(LegalEntity).where(LegalEntity.id == client_id))
    else:
        clients_r = await db.execute(select(LegalEntity).order_by(LegalEntity.legal_name))
    clients = clients_r.scalars().all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for client in clients:
        rfc = f" {client.rfc}" if client.rfc else ""
        sheet_name = f"{client.legal_name[:20]}{rfc}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        widths = [15, 20, 15, 15, 18, 18, 18, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # Title
        ws.merge_cells(f"A{row}:H{row}")
        cell_style(ws, row, 1, "ESTADO DE CUENTA", bold=True, color=COLOR_WHITE, bg=COLOR_HEADER, align="center")
        ws.row_dimensions[row].height = 28
        row += 1

        # Client info
        ws.merge_cells(f"A{row}:H{row}")
        cell_style(ws, row, 1, client.legal_name, bold=True, color=COLOR_WHITE, bg=COLOR_SUBHEADER)
        ws.row_dimensions[row].height = 20
        row += 1

        if client.trade_name:
            ws.merge_cells(f"A{row}:H{row}")
            cell_style(ws, row, 1, f"Nombre comercial: {client.trade_name}", color=COLOR_WHITE, bg=COLOR_SUBHEADER)
            row += 1

        if client.rfc:
            ws.merge_cells(f"A{row}:H{row}")
            cell_style(ws, row, 1, f"RFC: {client.rfc}", color=COLOR_WHITE, bg=COLOR_SUBHEADER)
            row += 1

        ws.merge_cells(f"A{row}:H{row}")
        cell_style(ws, row, 1, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", color="94A3B8", bg=COLOR_SUBHEADER)
        row += 2

        # Get invoices
        inv_r = await db.execute(
            select(Invoice).where(Invoice.client_id == client.id).order_by(Invoice.issue_date.desc())
        )
        invoices = inv_r.scalars().all()

        if invoices:
            # Invoices header
            ws.merge_cells(f"A{row}:H{row}")
            cell_style(ws, row, 1, "FACTURAS", bold=True, color=COLOR_WHITE, bg="1D4ED8", align="center")
            row += 1

            headers = ["Folio", "Fecha emisión", "Fecha vencimiento", "Total", "Cobrado", "Saldo", "Estado"]
            for col, h in enumerate(headers, 1):
                cell_style(ws, row, col, h, bold=True, color=COLOR_WHITE, bg="3B82F6", align="center")
            ws.row_dimensions[row].height = 18
            inv_table_start = row + 1
            row += 1

            for inv in invoices:
                status_text = {
                    "issued": "Emitida", "partial": "Parcial",
                    "paid": "Pagada", "overdue": "Vencida", "cancelled": "Cancelada"
                }.get(str(inv.status).replace("InvoiceStatus.", "").lower(), str(inv.status))

                balance = float(inv.balance or 0)
                bg = COLOR_LIGHT if invoices.index(inv) % 2 == 0 else COLOR_WHITE

                cell_style(ws, row, 1, inv.folio, bold=True, bg=bg)
                cell_style(ws, row, 2, inv.issue_date.strftime("%d/%m/%Y") if inv.issue_date else "—", bg=bg)
                cell_style(ws, row, 3, inv.due_date.strftime("%d/%m/%Y") if inv.due_date else "—", bg=bg)
                cell_style(ws, row, 4, float(inv.total or 0), bg=bg, align="right", number_format='$#,##0.00')
                cell_style(ws, row, 5, float(inv.paid_amount or 0), color="065F46", bg=bg, align="right", number_format='$#,##0.00')
                cell_style(ws, row, 6, balance, color=COLOR_AMBER if balance > 0 else "065F46", bg=bg, align="right", number_format='$#,##0.00')
                cell_style(ws, row, 7, status_text, bg=bg, align="center")
                row += 1

                # ── Payment history ───────────────────────────────────────
                payments_r = await db.execute(
                    select(InvoicePayment)
                    .where(InvoicePayment.invoice_id == inv.id)
                    .order_by(InvoicePayment.payment_date.asc())
                )
                payments = payments_r.scalars().all()

                if payments:
                    # Payment sub-header
                    cell_style(ws, row, 1, "", bg=COLOR_GRAY)
                    cell_style(ws, row, 2, "Fecha pago", bold=True, bg=COLOR_GRAY, color="475569")
                    cell_style(ws, row, 3, "Referencia", bold=True, bg=COLOR_GRAY, color="475569")
                    cell_style(ws, row, 4, "Notas", bold=True, bg=COLOR_GRAY, color="475569")
                    cell_style(ws, row, 5, "Monto", bold=True, bg=COLOR_GRAY, color="475569", align="right")
                    cell_style(ws, row, 6, "", bg=COLOR_GRAY)
                    cell_style(ws, row, 7, "", bg=COLOR_GRAY)
                    row += 1

                    for p in payments:
                        cell_style(ws, row, 1, "  ↳", bg=COLOR_GRAY, color="94A3B8")
                        cell_style(ws, row, 2, p.payment_date.strftime("%d/%m/%Y") if p.payment_date else "—", bg=COLOR_GRAY, color="334155")
                        cell_style(ws, row, 3, p.reference or "—", bg=COLOR_GRAY, color="334155")
                        cell_style(ws, row, 4, p.notes or "—", bg=COLOR_GRAY, color="334155")
                        cell_style(ws, row, 5, float(p.amount), color="065F46", bold=True, bg=COLOR_GRAY, align="right", number_format='$#,##0.00')
                        cell_style(ws, row, 6, "", bg=COLOR_GRAY)
                        cell_style(ws, row, 7, "", bg=COLOR_GRAY)
                        row += 1

            add_border(ws, inv_table_start, row - 1, 1, 7)

            # Totals
            row += 1
            total_f = sum(float(i.total or 0) for i in invoices)
            total_c = sum(float(i.paid_amount or 0) for i in invoices)
            total_s = sum(float(i.balance or 0) for i in invoices if str(i.status).replace("InvoiceStatus.", "").lower() not in ["paid", "cancelled"])

            cell_style(ws, row, 1, "TOTALES", bold=True, color=COLOR_WHITE, bg=COLOR_HEADER)
            cell_style(ws, row, 2, "", bg=COLOR_HEADER)
            cell_style(ws, row, 3, "", bg=COLOR_HEADER)
            cell_style(ws, row, 4, total_f, bold=True, color=COLOR_WHITE, bg=COLOR_HEADER, align="right", number_format='$#,##0.00')
            cell_style(ws, row, 5, total_c, bold=True, color=COLOR_WHITE, bg=COLOR_HEADER, align="right", number_format='$#,##0.00')
            cell_style(ws, row, 6, total_s, bold=True, color=COLOR_AMBER, bg=COLOR_HEADER, align="right", number_format='$#,##0.00')
            cell_style(ws, row, 7, "", bg=COLOR_HEADER)
            row += 2

        # Get quotes
        quote_r = await db.execute(
            select(Quote).where(Quote.client_id == client.id).order_by(Quote.issue_date.desc())
        )
        quotes = quote_r.scalars().all()

        if quotes:
            ws.merge_cells(f"A{row}:H{row}")
            cell_style(ws, row, 1, "COTIZACIONES", bold=True, color=COLOR_WHITE, bg="065F46", align="center")
            row += 1

            headers = ["Folio", "Fecha", "Atención", "Total", "Estado"]
            for col, h in enumerate(headers, 1):
                cell_style(ws, row, col, h, bold=True, color=COLOR_WHITE, bg="10B981", align="center")
            q_start = row + 1
            row += 1

            for q in quotes:
                status_text = {
                    "draft": "Borrador", "sent": "Enviada", "approved": "Aprobada",
                    "rejected": "Rechazada", "expired": "Expirada", "invoiced": "Facturada"
                }.get(str(q.status).replace("QuoteStatus.", "").lower(), str(q.status))

                bg = COLOR_LIGHT if quotes.index(q) % 2 == 0 else COLOR_WHITE
                cell_style(ws, row, 1, q.folio, bold=True, bg=bg)
                cell_style(ws, row, 2, q.issue_date.strftime("%d/%m/%Y") if q.issue_date else "—", bg=bg)
                cell_style(ws, row, 3, q.attention_name or "—", bg=bg)
                cell_style(ws, row, 4, float(q.total or 0), bg=bg, align="right", number_format='$#,##0.00')
                cell_style(ws, row, 5, status_text, bg=bg, align="center")
                row += 1

            add_border(ws, q_start, row - 1, 1, 5)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"estado_cuenta_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
